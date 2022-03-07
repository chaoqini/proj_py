#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
import numpy as np 
from matplotlib import pyplot as plt 

#class wbacbi:
#    def __init__(self,wbfile):
#        self.wbfile=wbfile
#        self.db=[]
#        self.nrowsPreData=5
#        self.unitIndex=1
#        self.wb2db(wbfile)

#    def wb2db(self,wbfile):
#        wb=openpyxl.load_workbook(wbfile)

#def g(x)->str:
#    for k,v in locals().items():
#        if v is x:
#            return k

def pt(x,display=1,fexit=0):
    print("dir is:\n",dir(x))
    print("type is:\n ",type(x))
    if display==1: print("variable is:\n",x)
    if fexit!=0: exit(0)
    return x



def wbread(wb_obj=None,num_sheet=0,wb_path=''):
    if wb_obj==None: wb=openpyxl.load_workbook(wb_path)
    else: wb=wb_obj
    wbs=wb.worksheets[num_sheet]
    maxc=wbs.max_column
    maxr=wbs.max_row
    db=[['' for i in range(1+maxr)]for i in range(1+maxc)]
    for x in range(maxc):
        for y in range(maxr):
            val=wbs.cell(column=x+1,row=y+1).value
            db[x+1][y+1]=(val if val is not None else '')
    return db

def wbwrite(db=[],num_sheet=0,wb_path='tmp.xlsx'):
    from openpyxl import Workbook
    wb=Workbook() 
    wbs=wb.worksheets[num_sheet]
    for x in range(1,len(db)):
        for y in range(1,len(db[x])):
            wbs.cell(column=x,row=y).value=db1[x][y]
    wb.save(wb_path)
    return wbs.title

wbp1='data1.xlsx'
#wbp2='test3.xlsx'
wb1=openpyxl.load_workbook(wbp1)
#wb2=openpyxl.load_workbook(wbp2)
db1=wbread(wb1,2)
#pt(len(db1[0]))
#wbs1=wb1.worksheets[0]
#pt(wbs1.max_row)
#pt(wbs1.title)
#pt(wb2.properties)
#x=wbread(wb_test1,0)
#pt(db1)
y=wbwrite(db1)
pt(y)
exit(0)


#wbfile='data1.xlsx'
#wbfile='/home/qc/download/2201-06-PM_VA05A_HTDR_Result_500hrs_20220216.xlsm'
#wbf_test1='test1.xlsx'
wbf_test1='test.xlsm'
wbfile1='data1.xlsx'
wb1=openpyxl.load_workbook(wbfile1)
wbfile2='test2.xlsx'
wb2=openpyxl.load_workbook(wbfile2)
wb_test1=openpyxl.load_workbook(wbf_test1)
#wb2=openpyxl.Workbook()
#wbn1=wb1.get_sheet_names()#获取sheet页
#wbs1=wb1.get_sheet_by_name(wbn1[-5])
#wbn2=wb2.get_sheet_names()
#wbs2=wb2.get_sheet_by_name(wbn2[0])
wb1s1=wb1.worksheets[2]
wb2s1=wb2.worksheets[0]
#wbts1=wb_test1.worksheets[-5]
wbts1=openpyxl.load_workbook(wbf_test1).worksheets[-5]
maxr=wb1s1.max_row    #最大行数
maxc=wb1s1.max_column  #最大列数
#pt(maxr)
#pt(maxc)
pt(wbts1.title)
#pt(wb1s1.title)
db=readsheet(wbf_test1,-5)
pt(db)
exit(0)
#x=wb1s1.cell(row=1,column=1).value
#db1=[]
db1=[['' for i in range(1+maxr)]for i in range(1+maxc)]
#val1=''
for x in range(maxc):
#for x in range(1):
    for y in range(maxr):
        val = wb1s1.cell(row=y+1,column=x+1).value
        db1[x+1][y+1] = (val if val is not None else '')

#pt(db1)

#with open('1.txt','w') as f:
#    for dc in db1:
#        f.write('\n')
#    for v in dc:
#            if v != '' :
##                print(v,"\r")
#                f.write(str(v)+' ')

#pt(len(db1))
#pt(len(db1[0]))
for c in range(3,maxc+1):
    data=db1[c][6:maxr+1]
    data=[i for i in data if i != '']
#    mean=np.mean(data)
#    db1[c].append(db1[c][48])
    if data!=[] :
        db1[c].append('')
        db1[c].append(db1[c][1])
        db1[c].append(db1[c][2])
        db1[c].append(np.mean(data))
        db1[c].append(np.std(data))
        db1[c].append(max(data))
        db1[c].append(min(data))

#pt(len(db1))
#pt(len(db1[4]))

wb2s2=wb2.worksheets[1]
i=0
for x in range(len(db1)):
    for y in range(len(db1[x])):
        wb2s2.cell(column=x+1,row=y+1).value=db1[x][y]
#pt(i)
#pt(wb2s2.title)
wb2.save('test2.xlsx')
