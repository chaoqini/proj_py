#!/usr/bin/env python
# -*- coding: utf-8 -*-

#import openpyxl
import numpy as np 
import pandas as pd
from matplotlib import pyplot as plt 


def pt(x,display=1,fexit=0):
    print("dir is:\n",dir(x))
    print("type is:\n ",type(x))
    if display==1: print("variable is:\n",x)
    if fexit!=0: exit(0)
    return x

##file='dt2.xlsx'
#file='test.xlsm'
file='dt1.xlsx'
dfn5=pd.read_excel(file,sheet_name=-5)
dfn6=pd.read_excel(file,sheet_name=-6)
#data=wb.head()
#pt(wb)
#pt(wb.shape)
#pt(wb.describe)
#pt(wb.info)
#wb.to_excel('tmp.xlsx')
#da=wb[46:]
#da=da.iloc[:,4:]
#da=wb.iloc[46:,5:]
#df2.loc[len(df2)]=['a','b','c']
#df3=df2.iloc[46:50,:]
#df3=df2.iloc[46:50]
#df2.append(df3)
#df4=pd.concat([df2,df3])
df1=dfn6.iloc[51:92,6:]
df2=dfn5.iloc[51:92,6:]
#pt(df5.head())
#for i in df5:
#df7=pd.DataFrame()
dfp=dfn5.iloc[48-2:51-1,7-1:]
dfd=df2
#pt(df6.shape[1])
#for i in df5.iteritems():
for x in range(df1.shape[0]):
    for y in range(df1.shape[1]):
#        print(df5.iloc[y,x])
#        print(df6.iloc[y,x])
        d1=df1.iloc[x,y]
        d2=df2.iloc[x,y]
        if d1!=0 :  dfd.iloc[x,y]=d2/d1
#        print(d3)
#    if n<=3 :  print(i)
#df7=df5.iloc[1,1]
#x=df7
#x=df6["Unnamed:13"].max()
#x=df6.max()
#pt(x)
#pt(dfd.max())
#pt(dfd.min())
#pt(dfd.mean())
#pt(dfd.std())
dfdmax=dfd.max().to_frame().T
dfdmin=dfd.min().to_frame().T
dfdmean=dfd.mean().to_frame().T
dfdstd=dfd.std().to_frame().T
#pt(dfdmax)
#dfa=pd.concat([dfd,dfp,dfdmax,dfdmin,dfdmean,dfdstd])
dfa=pd.concat([dfd,dfp,dfdmean,dfdstd,dfdmin,dfdmax])
#dfa=dfdmax
#dfa=dfdmax.swapaxes()
#dfa=dfdmax.swapaxes("index", "columns",copy=1)
#dfa=dfdmax.to_frame().T
#dfa=dfdmax.T
#dfa=dfdmax.transpose()
#df5.to_excel('tmp.xlsx',index=0)
#pt(dfa)
dfa.to_excel('tmp.xlsx')
#dfp.to_excel('tmp.xlsx')
#dfn5.to_excel('tmp.xlsx')

exit(0)


def wbread(wb_obj=None,num_sheet=0,wb_path=''):
    if wb_obj==None: wb=openpyxl.load_workbook(wb_path)
    else: wb=wb_obj
    wbs=wb.worksheets[num_sheet]
    maxc=wbs.max_column
    maxr=wbs.max_row
    db=[['' for i in range(1+maxr)]for i in range(1+maxc)]
    for x in range(maxc):
        for y in range(maxr):
            v=wbs.cell(column=x+1,row=y+1).value
            db[x][y]=(v if v is not None else '')
    return db

def wbwrite(db=[],num_sheet=0,wb_path='tmp.xlsx'):
    from openpyxl import Workbook
    wb=Workbook() 
    wbs=wb.worksheets[num_sheet]
    for x in range(len(db)):
        for y in range(len(db[x])):
            wbs.cell(column=x+1,row=y+1).value=db[x][y]
    wb.save(wb_path)
    return wbs.title


#data1='test1.xlsx'
#data1='test2.xlsx'
data1='dt2.xlsx'
wb1=openpyxl.load_workbook(data1)
db1=wbread(wb1,-6)
db2=wbread(wb1,-5)
#wbwrite(db1)
#wbwrite(db2)

d1=[c[53-1:-4] for c in db1[7-1:]]
d2=[c[53-1:-4] for c in db2[7-1:]]
d3=[['' for r in c] for c in d1]
#d3=(('' for r in c) for c in d1)
for c in range(len(d1)):
    for r in range(len(d1[c])):
        i1=d1[c][r];i2=d2[c][r]
        if i1!='' and i2!='' and float(i2)!=0:
            v=float(i1)/float(i2)
            v='%.3f%%'%v
#            v='%.4f'%v
            d3[c][r]=v
#            d3(c,r)=v


#for c in range(len(d3)):
#    for r in range(len(d3[c])):
#        i1=d1[c][r];i2=d2[c][r]
#        if i1!='' and i2!='' and float(i2)!=0:
#            v=float(i1)/float(i2)
#            v='%.3f%%'%v
#            d3[c][r]=v


#dd=[['aa']+i+['bb']+[min(i)]+[max(i)] for i in d3]
#dd=[[i]+['']*2+[min(i)]+[max(i)+[np.mean(i)]+[np.std(i)]] for i in d3]
#dd=[str(i)+[min(i)]+[max(i)+[np.mean(i)]+[np.std(i)]] for i in d3]
#dd=[[i,np.mean(i)] for i in d3]

#x=(i for i in d3[0] if i!='')
#y=np.mean(d3,axis=0)
nd=np.array(d3)
mean=np.mean(nd,axis=0)
x=mean
pt(x)
#dp=[c[47:53] for c in db1]

#da=[c+['']*2+['aa'] for c in dp]
#da=[i1+i2 for i1 in dp for i2 in d3]
#da=[ dp[6+i]+d3[i] for i in range(len(d3)) ]

#wbwrite(d3)
#wbwrite(dd)
#wbwrite(na)
exit(0)



