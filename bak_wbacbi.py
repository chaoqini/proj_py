#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
import numpy as np 
from matplotlib import pyplot as plt 

#class wbacbi:
#    def __init__(self,wbfile):
#        self.wbfile=wbfile
#        self.db=[]
#        self.nrowsPreData=5
#        self.unitIndex=1
#        self.wb2db(wbfile)

#    def wb2db(self,wbfile):
#        wb=openpyxl.load_workbook(wbfile)

#def g(x)->str:
#    for k,v in locals().items():
#        if v is x:
#            return k

def pt(x,display=1,fexit=0):
    print("dir is:\n",dir(x))
    print("type is:\n ",type(x))
    if display==1: print("variable is:\n",x)
    if fexit!=0: exit(0)
    return x

def wbread(wb_obj=None,num_sheet=0,wb_path=''):
    if wb_obj==None: wb=openpyxl.load_workbook(wb_path)
    else: wb=wb_obj
    wbs=wb.worksheets[num_sheet]
    maxc=wbs.max_column
    maxr=wbs.max_row
    db=[['' for i in range(1+maxr)]for i in range(1+maxc)]
    for x in range(maxc):
        for y in range(maxr):
            v=wbs.cell(column=x+1,row=y+1).value
            db[x][y]=(v if v is not None else '')
    return db

def wbwrite(db=[],num_sheet=0,wb_path='tmp.xlsx'):
    from openpyxl import Workbook
    wb=Workbook() 
    wbs=wb.worksheets[num_sheet]
    for x in range(len(db)):
        for y in range(len(db[x])):
            wbs.cell(column=x+1,row=y+1).value=db[x][y]
    wb.save(wb_path)
    return wbs.title


#data1='test1.xlsx'
#data1='test2.xlsx'
data1='dt2.xlsx'
wb1=openpyxl.load_workbook(data1)
db1=wbread(wb1,-6)
db2=wbread(wb1,-5)
#wbwrite(db1)
#wbwrite(db2)

d1=[c[53-1:-4] for c in db1[7-1:]]
d2=[c[53-1:-4] for c in db2[7-1:]]
d3=[['' for r in c] for c in d1]
#d3=(('' for r in c) for c in d1)
for c in range(len(d1)):
    for r in range(len(d1[c])):
        i1=d1[c][r];i2=d2[c][r]
        if i1!='' and i2!='' and float(i2)!=0:
            v=float(i1)/float(i2)
            v='%.3f%%'%v
#            v='%.4f'%v
            d3[c][r]=v
#            d3(c,r)=v


#for c in range(len(d3)):
#    for r in range(len(d3[c])):
#        i1=d1[c][r];i2=d2[c][r]
#        if i1!='' and i2!='' and float(i2)!=0:
#            v=float(i1)/float(i2)
#            v='%.3f%%'%v
#            d3[c][r]=v


#dd=[['aa']+i+['bb']+[min(i)]+[max(i)] for i in d3]
#dd=[[i]+['']*2+[min(i)]+[max(i)+[np.mean(i)]+[np.std(i)]] for i in d3]
#dd=[str(i)+[min(i)]+[max(i)+[np.mean(i)]+[np.std(i)]] for i in d3]
#dd=[[i,np.mean(i)] for i in d3]

#x=(i for i in d3[0] if i!='')
#y=np.mean(d3,axis=0)
nd=np.array(d3)
mean=np.mean(nd,axis=0)
x=mean
pt(x)
#dp=[c[47:53] for c in db1]

#da=[c+['']*2+['aa'] for c in dp]
#da=[i1+i2 for i1 in dp for i2 in d3]
#da=[ dp[6+i]+d3[i] for i in range(len(d3)) ]

#wbwrite(d3)
#wbwrite(dd)
#wbwrite(na)
exit(0)



